#!/usr/bin/env python
# coding: utf-8

import streamlit as st
import pandas as pd
import yagmail
import re
import tempfile
import os
import base64
from io import StringIO
from streamlit_quill import st_quill
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------
# CONFIG + BACKGROUND
# ---------------------------
st.set_page_config(page_title="Email Blaster", layout="wide", page_icon="📧")

# background helper (uses a local file if available)
def set_background(image_path="background.jpg"):
    if not os.path.exists(image_path):
        return
    with open(image_path, "rb") as f:
        data = f.read()
    encoded = base64.b64encode(data).decode()
    st.markdown(
        f"""
        <style>
        .stApp {{
            background-image: url("data:image/jpg;base64,{encoded}");
            background-size: cover;
            background-position: center;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

# call background (file name is background.jpg in same folder)
set_background("background.jpg")

# ---------------------------
# EXCEL LOGGING FUNCTION (CLEAN FORMAT)
# ---------------------------
def export_logs_excel(logs):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"email_logs_{ts}.xlsx"
    # Use tmp folder where available
    tmpdir = "/tmp" if os.path.exists("/tmp") else "."
    filepath = os.path.join(tmpdir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Logs"

    headers = ["Row", "Email", "Status", "Details", "Timestamp"]
    ws.append(headers)

    # Bold headers
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Insert log rows
    for row in logs:
        ws.append(row)

    # Borders and auto-width
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for r in ws.iter_rows():
        for cell in r:
            cell.border = thin

    # Auto column width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(filepath)
    return filepath, filename

# ---------------------------
# STYLES (full-page modern card)
# ---------------------------
st.markdown(
    """
    <style>
    /* Container layout */
    .page-center {
        display: flex;
        justify-content: center;
        align-items: flex-start;
        padding-top: 48px;
        padding-bottom: 48px;
    }

    .card {
        width: 980px;
        max-width: calc(100% - 48px);
        background: rgba(255,255,255,0.98);
        border-radius: 14px;
        padding: 28px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.12);
        border: 1px solid rgba(0,0,0,0.04);
        font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial;
    }

    .card-header {
        display:flex;
        align-items:center;
        justify-content:space-between;
        margin-bottom: 18px;
    }

    .title {
        font-size: 22px;
        font-weight: 700;
        color: #111827;
    }

    .subtitle {
        color: #6b7280;
        font-size: 14px;
        margin-top: 4px;
    }

    .section {
        margin-top: 18px;
        margin-bottom: 18px;
        padding: 16px;
        border-radius: 10px;
        background: #ffffff;
    }

    .section h3 {
        margin: 0 0 8px 0;
        font-size: 16px;
        color: #0f172a;
    }

    /* Buttons */
    .stButton>button {
        background: linear-gradient(90deg,#4b6ef6, #2f66f2) !important;
        color: white !important;
        border-radius: 8px !important;
        padding: 10px 14px !important;
        font-weight: 600;
        border: none !important;
    }

    /* Input rounding */
    .stTextInput>div>div>input, .stTextArea>div>div>textarea {
        border-radius: 8px !important;
    }

    /* Hide default menu/footers */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------------------------
# REMEMBER ME SYSTEM (Option B)
# ---------------------------
if "saved_email" not in st.session_state:
    st.session_state.saved_email = None
if "saved_pass" not in st.session_state:
    st.session_state.saved_pass = None
if "remember_email" not in st.session_state:
    st.session_state.remember_email = False
if "remember_pass_session" not in st.session_state:
    st.session_state.remember_pass_session = False

# load email from query params if present
query_params = st.experimental_get_query_params()
if "email" in query_params and st.session_state.saved_email is None:
    st.session_state.saved_email = query_params["email"][0]

# ---------------------------
# FIELD DETECTION (original logic)
# ---------------------------
FIELD_MAP = {
    "email": ["email", "mail", "e-mail", "emailaddress", "emailid"],
    "name": ["name", "fullname", "full name", "nama", "nama lengkap"],
    "company": ["company", "organization", "org", "perusahaan", "instansi"],
    "position": ["position", "jobtitle", "title", "jabatan", "role"]
}

def normalize(text):
    return re.sub(r"[^a-z0-9]", "", text.lower())

def detect_columns(df):
    detected = {}
    normalized_cols = {normalize(c): c for c in df.columns}

    for field, aliases in FIELD_MAP.items():

        # Hard match for email
        if field == "email":
            for norm, real in normalized_cols.items():
                if norm == "email":
                    detected[field] = real
                    break
            if field in detected:
                continue

        # Alias matching
        for alias in aliases:
            alias_norm = normalize(alias)
            for norm, real in normalized_cols.items():
                if alias_norm in norm or norm in alias_norm:
                    detected[field] = real
                    break
            if field in detected:
                break

    return detected

# ---------------------------
# QUILL SETUP (prevent reset)
# ---------------------------
if "body_html" not in st.session_state:
    st.session_state.body_html = ""
if "quill_initialized" not in st.session_state:
    st.session_state.quill_initialized = False

# ---------------------------
# LAYOUT: main centered card
# ---------------------------
st.markdown('<div class="page-center">', unsafe_allow_html=True)
st.markdown('<div class="card">', unsafe_allow_html=True)

# Header
st.markdown('<div class="card-header">', unsafe_allow_html=True)
st.markdown('<div><div class="title">📧 Email Blaster</div><div class="subtitle">Send personalised emails safely and easily</div></div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------
# SECTION 1 — Login (two-column)
# ---------------------------
st.markdown('<div class="section">', unsafe_allow_html=True)
st.markdown("### 1. Email Login")

col1, col2, col3 = st.columns([2,2,1])
with col1:
    st.session_state.remember_email = st.checkbox("Remember Email", value=st.session_state.remember_email)
with col2:
    st.session_state.remember_pass_session = st.checkbox("Remember App Password (session only)", value=st.session_state.remember_pass_session)
with col3:
    st.write("")  # spacer

# Email and password
email_user = st.text_input(
    "Your Email Address",
    placeholder="example@gmail.com",
    value=st.session_state.saved_email if st.session_state.saved_email else ""
)

email_pass = st.text_input(
    "App Password (NOT your regular password)",
    type="password",
    value=st.session_state.saved_pass if st.session_state.remember_pass_session else ""
)

st.info("For Gmail: create an App Password at https://myaccount.google.com/apppasswords")

# persist choices
if st.session_state.remember_email and email_user:
    st.session_state.saved_email = email_user
    # store to query params so it persists on reload in this browser
    st.experimental_set_query_params(email=email_user)

if st.session_state.remember_pass_session:
    st.session_state.saved_pass = email_pass
else:
    # clear saved_pass if unchecked
    st.session_state.saved_pass = None

st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------
# SECTION 2 — Email Details (big)
# ---------------------------
st.markdown('<div class="section">', unsafe_allow_html=True)
st.markdown("### 2. Email Details")

subject = st.text_input("Email Subject", "")

st.markdown("Email Body (Rich Text Editor)")
# Provide value only on first initialization
if not st.session_state.quill_initialized:
    editor_output = st_quill(
        value=st.session_state.body_html,
        html=True,
        placeholder="Write your email here... Use {name}, {company}, etc.",
        key="MAIN_EDITOR"
    )
    st.session_state.quill_initialized = True
else:
    editor_output = st_quill(
        html=True,
        placeholder="Write your email here... Use {name}, {company}, etc.",
        key="MAIN_EDITOR"
    )

if editor_output and editor_output != st.session_state.body_html:
    st.session_state.body_html = editor_output

st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------
# SECTION 3 — Upload recipients
# ---------------------------
st.markdown('<div class="section">', unsafe_allow_html=True)
st.markdown("### 3. Upload Recipients")

uploaded_excel = st.file_uploader("Upload .xlsx file (columns such as email, name, company, position)", type=["xlsx"])
df = None
detected_fields = {}

if uploaded_excel:
    try:
        df = pd.read_excel(uploaded_excel)
        detected_fields = detect_columns(df)
        st.success(f"Excel uploaded — {len(df)} rows found.")
        st.info(f"Detected fields: {detected_fields}")
    except Exception as e:
        st.error(f"Failed to read Excel: {e}")

st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------
# SECTION 4 — Preview
# ---------------------------
st.markdown('<div class="section">', unsafe_allow_html=True)
st.markdown("### 4. Preview")

if st.button("Show Preview"):
    if not st.session_state.body_html.strip():
        st.error("Email body is empty.")
    else:
        preview = st.session_state.body_html
        if df is not None and len(df) > 0:
            first = df.iloc[0]
            for field, col in detected_fields.items():
                preview = preview.replace(f"{{{field}}}", "" if pd.isna(first[col]) else str(first[col]))
        st.markdown("<div style='padding:12px;border-radius:8px;background:#fafafa;border:1px solid #eee'>", unsafe_allow_html=True)
        st.markdown(preview, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------
# SECTION 5 — Attachments
# ---------------------------
st.markdown('<div class="section">', unsafe_allow_html=True)
st.markdown("### 5. Attachments (optional)")

uploaded_files = st.file_uploader(
    "Upload attachments (pdf, jpg, png)",
    type=["pdf", "jpeg", "jpg", "png"],
    accept_multiple_files=True
)

st.markdown('</div>', unsafe_allow_html=True)

# ---------------------------
# SECTION 6 — Send & Logs
# ---------------------------
st.markdown('<div class="section">', unsafe_allow_html=True)
st.markdown("### 6. Send Emails")

if st.button("🚀 Send Now"):
    # Validations
    if df is None:
        st.error("Please upload an Excel file.")
        st.stop()
    if "email" not in detected_fields:
        st.error("No email column detected.")
        st.stop()
    if not email_user or not email_pass:
        st.error("Email + App Password required.")
        st.stop()
    if not subject.strip():
        st.error("Subject cannot be empty.")
        st.stop()
    if not st.session_state.body_html.strip():
        st.error("Email body cannot be empty.")
        st.stop()

    # Save attachments temporarily
    temp_paths = []
    try:
        for f in uploaded_files:
            ext = os.path.splitext(f.name)[1]
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
            tmp.write(f.read())
            tmp.close()
            temp_paths.append(tmp.name)
    except Exception as e:
        st.error("Attachment error.")
        st.stop()

    # Connect SMTP
    try:
        yag = yagmail.SMTP(email_user, email_pass)
    except Exception as e:
        st.error(f"SMTP Login Failed: {e}")
        st.stop()

    logs = []
    total = len(df)
    progress = st.progress(0)
    count = 0

    email_col = detected_fields["email"]

    for idx, row in df.iterrows():
        body = st.session_state.body_html
        for field, col in detected_fields.items():
            body = body.replace(f"{{{field}}}", "" if pd.isna(row[col]) else str(row[col]))

        # Split multiple emails in one cell
        raw_emails = re.split(r"[\/,; ]+", str(row[email_col]))
        emails = [e for e in raw_emails if "@" in e]

        if not emails:
            logs.append([idx, "", "NO_EMAIL", "SKIPPED", datetime.utcnow()])
            count += 1
            progress.progress(count / total)
            continue

        for email_addr in emails:
            try:
                yag.send(
                    to=email_addr,
                    subject=subject,
                    contents=body,
                    attachments=temp_paths
                )
                logs.append([idx, email_addr, "SENT", "OK", datetime.utcnow()])
            except Exception as err:
                logs.append([idx, email_addr, "FAILED", str(err), datetime.utcnow()])

        count += 1
        progress.progress(count / total)

    # export excel log and offer download
    excel_path, excel_name = export_logs_excel(logs)
    st.success("All emails processed!")

    with open(excel_path, "rb") as f:
        st.download_button("📥 Download Logs (Excel)", f, file_name=excel_name)

    # cleanup
    for p in temp_paths:
        try:
            os.unlink(p)
        except:
            pass

st.markdown('</div>', unsafe_allow_html=True)

# close card and page-center
st.markdown('</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)
